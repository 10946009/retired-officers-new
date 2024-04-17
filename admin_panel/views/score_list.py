from django.shortcuts import redirect, render
from myapp.models import Activity, Score
import openpyxl
from django.http import HttpResponse
from admin_panel.forms import UploadExcelForm
from openpyxl import Workbook
from django.db import transaction
from django.contrib import messages
from django.contrib.auth.decorators import permission_required
# import custom User model
from user.models import Student, User,ActivityStudents
from decimal import Decimal


def get_student_score(student, activity):
    '''
    Get student score
    '''
    try:
        score = student.score_set.get(activity=activity)
        return score.score1, score.score2, score.score3
    except Score.DoesNotExist:
        print("Score does not exist for this student and activity.")
        return 0, 0, 0


@permission_required('myapp.view_activity', login_url='/403')
def score_list(request, activity_id):
    '''
    顯示學生成績列表
    '''
    activity = (
        Activity.objects.select_related("score_label")
        .prefetch_related("student")
        .get(id=activity_id)
    )
    # get have checked_number student
    score_label = activity.score_label
    students  = ActivityStudents.get_is_checked_student(activity_id)
    print(students)
    
    # get all students with their scores

    student_with_scores = []
    for student in students:
        student = student.student
        score1, score2, score3 = get_student_score(student, activity)
        total_score = sum([score1, score2, score3])

        student_with_scores.append(
            {
                "checked_number": student.get_checked_number(activity_id),
                "id": student.id,
                "name": student.user.username,
                "email": student.user.email,
                "score1": score1,
                "score2": score2,
                "score3": score3,
                "total": total_score,
            }
        )
    # student_with_scores sort for total
    student_with_scores = sorted(student_with_scores, key=lambda x: x["total"], reverse=True)
    return render(
        request,
        "score_list.html",
        {
            "student_with_scores": student_with_scores,
            "score_label": score_label,
            "activity": activity,
        },
    )

@permission_required('myapp.view_activity', login_url='/403')
def export_score_sample(request, activity_id):
    '''
    匯出成績範本
    '''

    # 創建一個 Excel 工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "sample"

    # 向工作表添加一些數據
    # Rest of the code...

    # find all students for this activity
    activity = Activity.objects.get(id=activity_id)
    # 取得該活動而且已經確認OK的學生資料
    students  = ActivityStudents.get_is_checked_student(activity_id)

    ws.append(
        [
            "ID",
            "報名編號",
            "報名證號",
            "姓名",
            "email",
            activity.score_label.label1,
            activity.score_label.label2,
            activity.score_label.label3,
        ]
    )
    for student in students:
        join_number = student.join_number
        checked_number = student.checked_number
        student = student.student
        ws.append([student.id,join_number,checked_number, student.user.username, student.user.email, 0, 0, 0])

    # 將工作簿保存到響應中
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=import_sample.xlsx"
    wb.save(response)

    return response

@permission_required('myapp.view_activity', login_url='/403')
def upload_and_read_excel(request, activity_id):
    '''
    匯入成績
    '''
    if request.method == "POST":
        form = UploadExcelForm(request.POST, request.FILES)

        if form.is_valid():
            excel_file = request.FILES["excel_file"]
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            print("111")
            # 初始化數據列表並跳過首行
            data = [row for row in ws.iter_rows(values_only=True) if row[0] != "ID"]
            print(data)
            activity = Activity.objects.get(id=activity_id)
            student_ids = [row[0] for row in data]

            # 先取得有確認OK的學生
            students = ActivityStudents.get_is_checked_student(activity_id)
            scores = Score.objects.filter(
                activity=activity, student__id__in=student_ids
            ).prefetch_related("student")

            scores_to_update = []
            scores_to_create = []

            with transaction.atomic():
                for row in data:
                    student_id, score1, score2, score3 = row[0], Decimal(row[5]), Decimal(row[6]), Decimal(row[7])
                    student = next((st.student for st in students if st.student.id == student_id), None)
                    # check if student exists
                    if student is None:
                        messages.error(
                            request, f"匯入失敗！ {row[0]} {row[1]} 學生不存在。"
                        )
                        return redirect("score_list", activity_id=activity_id)
                    # check if score is valid
                    score = next(
                        (sc for sc in scores if sc.student.id == student_id), None
                    )

                    if score:
                        score.score1 = score1
                        score.score2 = score2
                        score.score3 = score3
                        scores_to_update.append(score)
                    else:
                        scores_to_create.append(
                            Score(
                                student=student,
                                activity=activity,
                                score1=score1,
                                score2=score2,
                                score3=score3,
                            )
                        )

                # 使用 bulk_update 和 bulk_create 提高效率
                Score.objects.bulk_update(
                    scores_to_update, ["score1", "score2", "score3"]
                )
                Score.objects.bulk_create(scores_to_create)
                messages.success(request, "匯入成功！")

    # refresh the page
    return redirect("score_list", activity_id=activity_id)
