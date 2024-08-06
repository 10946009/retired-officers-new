from django.shortcuts import render
from django.contrib.auth.decorators import permission_required
from django.shortcuts import redirect, render
from myapp.models import Activity, Score, VoluntarySchool
import openpyxl
from django.http import HttpResponse
from admin_panel.forms import UploadExcelForm
from openpyxl import Workbook
from django.db import transaction
from django.contrib import messages
from django.contrib.auth.decorators import permission_required
# import custom User model
from user.models import ActivityStudents
from decimal import Decimal

@permission_required("myapp.view_activity", login_url="/403")
def voluntary_school(request, activity_id):
    voluntary_school = VoluntarySchool.objects.all()

    content = {"voluntary_school": voluntary_school, "activity_id": activity_id}
    return render(request, "voluntary_school.html", content)


@permission_required('myapp.view_activity', login_url='/403')
def export_voluntary_sample(request, activity_id):
    '''
    匯入學校志願範本
    '''

    # 創建一個 Excel 工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "sample"

    ws.append(
        [
            "校名",
            "學制",
            "類別",
            "系組",
            "核定名額",
            "志願代碼",
        ]
    )
    ws.append(["(請記得刪除範例資料)中國文化大學","在職專班","商管類","國際企業管理學系","2","001"])

    # 將工作簿保存到響應中
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=import_school_sample.xlsx"
    wb.save(response)

    return response

@permission_required('myapp.view_activity', login_url='/403')
def upload_voluntary_excel(request, activity_id):
    '''
    匯入學校志願
    '''
    if request.method == "POST":
        form = UploadExcelForm(request.POST, request.FILES)

        if form.is_valid():
            excel_file = request.FILES["excel_file"]
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            # 初始化數據列表並跳過首行
            data = [row for row in ws.iter_rows(values_only=True) if row[0] != "校名"]
            print(data)
            activity = Activity.objects.get(id=activity_id)
            with transaction.atomic():
                # 遍歷數據列表
                for row in data:
                    voluntary_school = VoluntarySchool.objects.create(
                        activity = activity,
                        name = row[0],
                        academic = row[1],
                        category = row[2],
                        department = row[3],
                        quota = row[4],
                        code = row[5]
                    )
                    voluntary_school.save()              

    # refresh the page
    return redirect("voluntary_school", activity_id=activity_id)
